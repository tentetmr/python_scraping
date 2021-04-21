
import requests
from bs4 import BeautifulSoup
import openpyxl


URL = "https://pdt.r-agent.com/"
LOGIN = "pdt/app/pdt_login_view"
LOGIN_ROUTE = "pdt/app/pdt_login?PDT31A"
USER = ""
PASS = ""

s = requests.session()
response = s.get(URL + LOGIN)
bs = BeautifulSoup(response.content, "html.parser")
authenticity = bs.find(attrs={'name': 'sn'}).get('value')
cookie = response.cookies
wb = openpyxl.Workbook()
wa = wb.active

info = {
    "sn": authenticity,
    "updateCheckFlag": "false",
    "accountId": USER,
    "password": PASS,
    "autoLoginFlag": "true",
    "complete": "true"
}

res = s.post(URL + LOGIN_ROUTE, data=info, cookies=cookie)
soup = BeautifulSoup(res.content, "html.parser")

# print(soup)

companyNames = soup.find_all("p", class_="mod-jobList-subHeading")

attrs = {"data-detail": "salary"}
companySalaries = soup.find_all("div", **attrs)

attrs2 = {"data-detail": "employeeNumber"}
companyEmployees = soup.find_all("div", **attrs2)

companyName = []
companySalary = []
companyEmployee = []

for elem in companyNames:
    companyName.append(elem.text)
for elem2 in companySalaries:
    companySalary.append(elem2.text)
for elem3 in companyEmployees:
    companyEmployee.append(elem3.text)

for i in range(0, len(companyName)):
    wa.cell(row=i+1, column=1, value=companyName[i-1])
    wa.cell(row=i+1, column=2, value=companySalary[i-1])
    wa.cell(row=i+1, column=3, value=companyEmployee[i-1])


wb.save(r'C:\Users\makal\OneDrive\デスクトップ\sample.xlsx')
